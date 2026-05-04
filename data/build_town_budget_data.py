#!/usr/bin/env python3
"""Build the structured FY27 budget JSON consumed by town-budget.html.

Sources combined:

1. ``data/FY27_Proposed_Budget_No_Override.txt`` -- whitespace dump of
   the FY27 Proposed Budget book PDF. Provides function-level vote
   totals, department headers, FY25 budget + FY25 actual columns, and
   the original "TOTAL GENERAL FUND ACCOUNTS" grand total.

2. ``data/budget_source/FY27-Proposed-Budget-vs-FY26-w-Acct-Details.xlsx``
   -- the Town's Munis-account-level worksheet (FY26 vs FY27, ~600 rows).
   When present, replaces the rolled-up "Salaries / Expense" lines from
   the budget book with one row per Munis account code (individual
   salaried positions, supply categories, etc.). Excel covers town-side
   general-fund departments only -- not enterprise funds, not schools.

3. ``data/peer_schedule_a_expenditures.csv`` -- DOR Schedule A FY02-FY24
   for sparkline history on function rows.

4. ``data/schools/sc-meetings-fy26/agenda-and-materials-2-5-2026-fy27-budget-packet.txt``
   -- per-school cost-center totals + Munis-level lines.

Output: data/town_budget_FY27.json + data/town_budget_FY27_lookup.json

Usage:
  python3 data/build_town_budget_data.py
"""
from __future__ import annotations
import csv
import json
import re
from pathlib import Path
from typing import Optional

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"

# Pattern matches a line like:
#   "TOTAL BUDGETS    115,368,206   111,991,311   119,479,480   122,762,030   3,282,550   2.75%"
# Six numeric columns, last is a percentage.
_NUM = r"[\d,]+|-"
_PCT = r"-?[\d.]+%"
_CHANGE = r"(?:\([\d,]+\)|[\d,]+|-)"
_GRAND_TOTAL_RE = re.compile(
    rf"^\s*TOTAL BUDGETS\s+({_NUM})\s+({_NUM})\s+({_NUM})\s+({_NUM})\s+({_CHANGE})\s+({_PCT})",
    re.MULTILINE,
)

_GENERAL_FUND_TOTAL_RE = re.compile(
    rf"^\s*TOTAL GENERAL FUND ACCOUNTS\s+({_NUM})\s+({_NUM})\s+({_NUM})\s+({_NUM})\s+({_CHANGE})\s+({_PCT})",
    re.MULTILINE,
)

# Map "VOTE TOTAL <CAPS DESCRIPTION>" → slug.
_FUNCTION_DESC_TO_SLUG = {
    "VOTE TOTAL GENERAL GOVERNMENT":            "general_government",
    "VOTE TOTAL PUBLIC SAFETY":                 "public_safety",
    "VOTE TOTAL SCHOOLS":                       "schools",
    "VOTE TOTAL PUBLIC WORKS AND FACILITIES":   "public_works",
    "VOTE TOTAL HUMAN SERVICES":                "human_services",
    "VOTE TOTAL CULTURE AND RECREATION":        "culture_recreation",
    "VOTE TOTAL OTHER GENERAL GOVERNMENT":      "other_general_government",
    "VOTE TOTAL SEWER ENTERPRISE FUND":         "sewer_enterprise",
    "VOTE TOTAL WATER ENTERPRISE FUND":         "water_enterprise",
    "VOTE TOTAL HARBOR ENTERPRISE FUND":        "harbor_enterprise",
}

_VOTE_TOTAL_RE = re.compile(
    rf"^\s*(VOTE TOTAL [A-Z& ]+?)\s+({_NUM})\s+({_NUM})\s+({_NUM})\s+({_NUM})\s+({_CHANGE})\s+({_PCT})",
    re.MULTILINE,
)

# A line item: "62 Salaries  4,644,044  4,608,022  4,730,006  4,988,616  258,610  5.47%"
_LINE_ITEM_RE = re.compile(
    rf"^\s*(\d+)\s+([A-Za-z][A-Za-z &/-]+?)\s+({_NUM})\s+({_NUM})\s+({_NUM})\s+({_NUM})\s+({_CHANGE})\s+({_PCT})\s*$"
)

# Department header: caps-or-mixed line ending in known suffixes.
_DEPT_HEADER_RE = re.compile(
    r"^\s+([A-Z][A-Za-z][A-Za-z &/'-]+?(?: Department| Inspector| Inspection| Fund| Reserve| Inspectional Services| Counsel| Clerk| Board| Committee| Department| Commission)?)\s*$"
)

# Map known department display names → slug.
_DEPT_DISPLAY_TO_SLUG = {
    "Police Department": "police",
    "Fire Department": "fire",
    "Building Inspection Department": "building_inspection",
    "Sealer of Weights and Measures": "sealer_weights_measures",
    "Animal Inspector": "animal_inspector",
    "Emergency Management Department": "emergency_management",
    "Harbor & Waterways Board": "harbor_waterways_board",
    "Police": "police",
    "Fire": "fire",
    "Moderator": "moderator",
    "Select Board": "select_board",
    "Finance Committee": "finance_committee",
    "Reserve Fund": "reserve_fund",
    "Finance Department": "finance",
    "Assessor": "assessor",
    "Town Counsel": "town_counsel",
    "Parking Clerk": "parking_clerk",
    "Town Clerk": "town_clerk",
    "Election and Registration Department": "election_registration",
    "Planning Board": "planning_board",
    "Public Buildings Department": "public_buildings",
    "Human Resources Department": "human_resources",
    "Comm Dev & Planning Department": "community_development",
    "Public Works Department": "public_works_dept",
    "Public Works (Highway, Tree, Drains)": "public_works_ops",
    "Trash and Recycling": "trash_recycling",
    "Waste Collection": "waste_collection",
    "NEW Curbside Collection": "curbside_collection",
    "Cemetery Department": "cemetery",
    "Health Department": "health",
    "Council on Aging": "council_on_aging",
    "Veterans Benefits": "veterans_benefits",
    "Abbot Public Library": "library",
    "Recreation and Park Department": "rec_park",
    "Memorial & Veterans Day": "memorial_veterans_day",
    "School Department": "schools_dept_wrapper",
    "Maturing Bonds and Interest": "debt_service",
    "Other General Government": "other_general_government_dept",
    "Sewer": "sewer",
    "Water": "water",
    "Light": "light",
    "Harbor": "harbor",
}

# Function vote-totals in the order they appear in the doc, used to
# attribute departments encountered between two VOTE TOTALs to a function.
_FUNCTION_ORDER = [
    "general_government",
    "public_safety",
    "schools",
    "public_works",
    "human_services",
    "culture_recreation",
    "other_general_government",
    "sewer_enterprise",
    "water_enterprise",
    "harbor_enterprise",
]


def _parse_int(s: str) -> int:
    s = s.strip().replace(",", "")
    if s in {"", "-"}:
        return 0
    if s.startswith("(") and s.endswith(")"):
        return -int(s[1:-1])
    return int(s)


def _parse_pct(s: str) -> float:
    s = s.strip().rstrip("%")
    if s in {"", "-"}:
        return 0.0
    return float(s) / 100.0


def _make_grand_total_row(slug: str, descr: str, m: re.Match) -> dict:
    return {
        "id": slug,
        "level": "grand_total",
        "parent_id": None,
        "function": None,
        "department": None,
        "description": descr,
        "spend_type": None,
        "fy25_budget": _parse_int(m.group(1)),
        "fy25_actual": _parse_int(m.group(2)),
        "fy26_budget": _parse_int(m.group(3)),
        "fy27_proposed": _parse_int(m.group(4)),
        "change_dollars": _parse_int(m.group(5)),
        "change_pct": _parse_pct(m.group(6)),
    }


def _make_function_row(slug: str, descr: str, m: re.Match) -> dict:
    return {
        "id": slug,
        "level": "function",
        "parent_id": None,
        "function": slug,
        "department": None,
        "description": descr,
        "spend_type": None,
        "fy25_budget": _parse_int(m.group(2)),
        "fy25_actual": _parse_int(m.group(3)),
        "fy26_budget": _parse_int(m.group(4)),
        "fy27_proposed": _parse_int(m.group(5)),
        "change_dollars": _parse_int(m.group(6)),
        "change_pct": _parse_pct(m.group(7)),
    }


def classify_spend_type(descr: str, munis_obj: Optional[str] = None) -> str:
    """Bucket a line-item description into a spend type.

    Buckets: salaries | expense | officials_expense | benefits | debt
    | transfer | utility | reserve | other.

    If ``munis_obj`` is provided (Munis object code), use the code's
    leading digits to classify first. Falls back to description matching.
    """
    if munis_obj:
        # 517xxx: benefits umbrella (retirement, insurance, medicare,
        # unemployment, workers comp, flex spending). Some 517 lines
        # describe payroll-style payouts; treat sick-bonus-retirement
        # as salaries because that's how it's pooled in the town book.
        if munis_obj.startswith("517"):
            d = descr.lower()
            if "sick bonus retirement" in d:
                return "salaries"
            return "benefits"
        # 51xxxx (other): salaries, OT, longevity, sick, stipends,
        # incentive, holiday, night-diff -- all payroll.
        if munis_obj.startswith("51"):
            return "salaries"
        # 521xxx: utilities (electricity, gas, oil, water/sewer charges)
        if munis_obj.startswith("521") or munis_obj.startswith("523"):
            return "utility"
        # 591xxx / 592xxx: long-term & short-term debt
        if munis_obj.startswith("591") or munis_obj.startswith("592"):
            return "debt"
        # 596xxx: interfund transfers
        if munis_obj.startswith("596"):
            return "transfer"
        # 572500: officials expense (moderator, elected positions)
        if munis_obj == "572500":
            return "officials_expense"
        # 579900: reserve fund transfer account
        if munis_obj == "579900":
            return "reserve"
        # Everything else under 5xxxxx is operating expense (supplies,
        # services, dues, repairs, travel, etc.)
        if munis_obj.startswith("5"):
            return "expense"

    d = descr.lower()
    if "salaries" in d or "salary reserve" in d:
        return "salaries"
    if "officials expense" in d:
        return "officials_expense"
    if "expense" in d:
        return "expense"
    if "benefits" in d or "retirement" in d or "insurance" in d \
       or "medicare" in d or "unemployment" in d \
       or "post employment" in d or "workers compensation" in d \
       or "flex spending" in d:
        return "benefits"
    if "maturing debt" in d or "interest" in d or "bonds" in d:
        return "debt"
    if "transfer" in d or "stabilization" in d:
        return "transfer"
    if "lighting" in d or "utility reserve" in d or "energy reserve" in d:
        return "utility"
    if "reserves" in d or "reserve fund" in d:
        return "reserve"
    return "other"


def parse_budget_book(text: str) -> list[dict]:
    """Parse the FY27 budget book text into a flat list of rows.

    Rows have fields: id, level, parent_id, function, department,
    description, spend_type, fy25_budget, fy25_actual, fy26_budget,
    fy27_proposed, change_dollars, change_pct.
    Line-level rows additionally have source_ref.
    """
    rows: list[dict] = []

    # Pass 1: extract grand totals + function totals.
    m = _GENERAL_FUND_TOTAL_RE.search(text)
    if m:
        rows.append(_make_grand_total_row("total_general_fund",
                                          "TOTAL GENERAL FUND ACCOUNTS", m))
    for m in _VOTE_TOTAL_RE.finditer(text):
        descr = m.group(1).strip()
        slug = _FUNCTION_DESC_TO_SLUG.get(descr)
        if slug:
            rows.append(_make_function_row(slug, descr, m))
    m = _GRAND_TOTAL_RE.search(text)
    if m:
        rows.append(_make_grand_total_row("total_budgets", "TOTAL BUDGETS", m))

    # Pass 2: line items and department headers.
    current_function_idx = 0
    current_dept_slug: Optional[str] = None
    current_dept_descr: Optional[str] = None
    dept_subtotals: dict[str, dict] = {}

    lines = text.splitlines()
    for raw in lines:
        if "VOTE TOTAL" in raw and current_function_idx < len(_FUNCTION_ORDER):
            current_function_idx += 1
            current_dept_slug = None
            continue

        line_m = _LINE_ITEM_RE.match(raw)
        if line_m:
            num, descr = line_m.group(1), line_m.group(2).strip()
            fy25_b = _parse_int(line_m.group(3))
            fy25_a = _parse_int(line_m.group(4))
            fy26_b = _parse_int(line_m.group(5))
            fy27_p = _parse_int(line_m.group(6))
            chg_d = _parse_int(line_m.group(7))
            chg_p = _parse_pct(line_m.group(8))
            fn_slug = (_FUNCTION_ORDER[current_function_idx]
                       if current_function_idx < len(_FUNCTION_ORDER) else None)
            line_id = f"line_{num}"
            rows.append({
                "id": line_id,
                "level": "line",
                "parent_id": current_dept_slug,
                "function": fn_slug,
                "department": current_dept_slug,
                "description": descr,
                "spend_type": classify_spend_type(descr),
                "fy25_budget": fy25_b,
                "fy25_actual": fy25_a,
                "fy26_budget": fy26_b,
                "fy27_proposed": fy27_p,
                "change_dollars": chg_d,
                "change_pct": chg_p,
                "source_ref": {
                    "doc": "fy27_budget_book",
                    "line_item": int(num),
                },
            })
            if current_dept_slug:
                d = dept_subtotals.setdefault(current_dept_slug, {
                    "id": current_dept_slug,
                    "level": "department",
                    "parent_id": fn_slug,
                    "function": fn_slug,
                    "department": current_dept_slug,
                    "description": current_dept_descr,
                    "spend_type": None,
                    "fy25_budget": 0, "fy25_actual": 0,
                    "fy26_budget": 0, "fy27_proposed": 0,
                    "change_dollars": 0,
                })
                d["fy25_budget"] += fy25_b
                d["fy25_actual"] += fy25_a
                d["fy26_budget"] += fy26_b
                d["fy27_proposed"] += fy27_p
                d["change_dollars"] += chg_d
            continue

        # Check explicit display-name map first (handles names with parens,
        # "NEW" prefixes, and other patterns the regex won't catch).
        stripped = raw.strip()
        if stripped in _DEPT_DISPLAY_TO_SLUG:
            current_dept_slug = _DEPT_DISPLAY_TO_SLUG[stripped]
            current_dept_descr = stripped
            continue

        dept_m = _DEPT_HEADER_RE.match(raw)
        if dept_m:
            dept_descr = dept_m.group(1).strip()
            dept_slug = _DEPT_DISPLAY_TO_SLUG.get(dept_descr)
            if dept_slug is None:
                dept_slug = re.sub(r"[^a-z0-9]+", "_", dept_descr.lower()).strip("_")
            current_dept_slug = dept_slug
            current_dept_descr = dept_descr

    # Compute change_pct for dept subtotals and append.
    for d in dept_subtotals.values():
        prior = d["fy26_budget"]
        d["change_pct"] = (d["change_dollars"] / prior) if prior else 0.0
        rows.append(d)

    # Drop department rows that are completely zeroed in both FY26 and FY27.
    # This removes eliminated departments (e.g. "Engineer" was folded into
    # Public Works in FY25; its FY27 budget book entry is $0 in both years).
    rows = [
        r for r in rows
        if not (
            r["level"] == "department"
            and (r.get("fy26_budget") or 0) == 0
            and (r.get("fy27_proposed") or 0) == 0
        )
    ]

    return rows


# Schedule A function-bucket → our function slug.
# `other_general_government` rolls up four Schedule A buckets that the FY27
# budget book combines under "Other General Government" + "Debt Service" lines.
SCHEDULE_A_FUNCTION_MAPPING = {
    "general_government": ["general_government"],
    "public_safety":      ["public_safety"],
    "schools":            ["education"],
    "public_works":       ["public_works"],
    "human_services":     ["human_services"],
    "culture_recreation": ["culture_recreation"],
    "other_general_government":
        ["fixed_costs", "intergov_assessments", "other", "debt_service"],
}


def load_schedule_a_history() -> dict[str, dict[str, int]]:
    """Read peer_schedule_a_expenditures.csv, filter to Marblehead, return
    {function_slug: {fyXX_actual: int}}.
    Combines (sums) buckets per SCHEDULE_A_FUNCTION_MAPPING for our slugs.
    """
    p = DATA / "peer_schedule_a_expenditures.csv"
    by_year: dict[int, dict[str, int]] = {}
    with p.open(newline="") as f:
        for row in csv.DictReader(f):
            if row["municipality"] != "Marblehead":
                continue
            fy = int(row["fiscal_year"])
            year_buckets: dict[str, int] = {}
            for col in ["general_government", "public_safety", "education",
                        "public_works", "human_services", "culture_recreation",
                        "fixed_costs", "intergov_assessments", "other",
                        "debt_service"]:
                val = row.get(col, "").replace(",", "").strip()
                year_buckets[col] = int(val) if val and val != "-" else 0
            by_year[fy] = year_buckets

    history: dict[str, dict[str, int]] = {}
    for fn_slug, schedule_a_buckets in SCHEDULE_A_FUNCTION_MAPPING.items():
        history[fn_slug] = {}
        for fy, buckets in by_year.items():
            total = sum(buckets.get(b, 0) for b in schedule_a_buckets)
            # Use 2-digit year suffix: 2002 → fy02, 2024 → fy24.
            yy = fy % 100
            history[fn_slug][f"fy{yy:02d}_actual"] = total
    return history


def attach_function_history(rows: list[dict]) -> None:
    """Attach `history` and `cagr_22yr` to each function-level row that has
    Schedule A coverage. Mutates rows in place."""
    history_by_fn = load_schedule_a_history()
    for r in rows:
        if r["level"] != "function":
            continue
        fn_history = history_by_fn.get(r["id"])
        if not fn_history:
            continue  # enterprise funds, schools_dept_wrapper, etc.
        r["history"] = fn_history
        # Compute CAGR FY02 → FY24 (22-year span).
        start = fn_history.get("fy02_actual")
        end = fn_history.get("fy24_actual")
        if start and end and start > 0:
            r["cagr_22yr"] = (end / start) ** (1 / 22) - 1


# Munis line item pattern. Matches lines like:
#   PROFESSIONAL SALARIES - SPED    1212021   510100 500   $   107,110   $   110,323   $   3,213   3.00%
#   Instructional Assistants - General Ed   2233013   510302   100   $   168,687   $   177,402   $   8,716   4.91%
# Description: one or more words (upper or mixed case), may include -/&.'
# ORG: 7 digits, OBJ: 6 digits, PROJ: 1-3 digits
# FY26 and FY27 amounts: $ NNN,NNN with optional spaces
# Change $: may be paren-wrapped (negative), may be missing FY26 for new items
# Change %: numeric with %, may be negative or absent
_MUNIS_LINE_RE = re.compile(
    r"^\s*"
    r"([A-Za-z][A-Za-z0-9 &/\-.'*]+?)\s+"    # description (upper or mixed, may trail *)
    r"(\d{7})\s+"                              # ORG (7 digits)
    r"(\d{6})\s+"                              # OBJ (6 digits)
    r"(\d{1,3})\s+"                            # PROJ (1-3 digits)
    r"\$\s*([\d,]+)\s+"                        # FY26 $ amount
    r"\$\s*([\d,]+)\s+"                        # FY27 $ amount
    r"\$\s*(\([\d,]+\)|[\d,]+)\s*"            # $ Change (may be parens-negative)
    r"(-?[\d.]+%)?",                           # % Change (optional)
    re.MULTILINE,
)

# Lines to skip even if they match the pattern.
_SKIP_DESCR_RE = re.compile(
    r"^(Sub Total|TOTAL|.*TOTAL\s*:|.*SERVICES TOTAL|.*PROGRAMS TOTAL"
    r"|.*ADMINSTR Total|.*TOTAL$|Level Fund|ADDITIONS|NEW REQUESTS"
    r"|Maintenance Sub Total)",
    re.IGNORECASE,
)


def parse_school_packet() -> list[dict]:
    """Parse the FY27 school budget packet for per-school cost-center totals
    and Munis-level sub-line items.

    Returns rows for the schools that could be parsed cleanly. If the packet
    format prevents extraction (different layout in a future year), returns
    an empty list -- the UI will fall back to showing Schools as one $47.6M
    lump matching the town book.

    The returned list contains:
    - 6 department-level rows (one per school) with level="department"
    - N line-level rows (one per Munis line item) with level="line"
    """
    p = (DATA / "schools" / "sc-meetings-fy26"
         / "agenda-and-materials-2-5-2026-fy27-budget-packet.txt")
    if not p.exists():
        return []
    try:
        text = p.read_text()
    except OSError:
        return []

    rows: list[dict] = []

    # Schools to extract: (slug, header pattern substring, display name)
    SCHOOLS = [
        ("school_brown",    "Marblehead Public Schools - Brown Elementary School",
         "Brown Elementary School"),
        ("school_glover",   "Marblehead Public Schools - Glover Elementary School",
         "Glover Elementary School"),
        ("school_village",  "Marblehead Public Schools - Village Elementary School",
         "Village Elementary School"),
        ("school_middle",   "Marblehead Public Schools - Veterans Middle School",
         "Veterans Middle School"),
        ("school_high",     "Marblehead Public Schools - Marblehead High School",
         "Marblehead High School"),
        ("school_athletics", "Marblehead Public Schools - Athletics",
         "Athletics"),
    ]

    # Pattern matching the TOTAL line format (various whitespace layouts):
    #   TOTAL :  $  5,481,505   $  5,554,252   $  72,747   1.33%
    #   TOTAL :  $  646,395     $  660,030     $  13,635    2.11%
    #   TOTAL :  $  4,472,968   $  4,437,815   $  (35,152) $  (0)   <- no % sign
    # The percentage column is optional/inconsistent; compute pct from dollars.
    total_re = re.compile(
        r"TOTAL\s*:\s+\$\s*([\d,]+)\s+\$\s*([\d,]+)\s+\$\s*\(?([\d,]+)\)?"
    )

    # Only search the first half of the document (level-fund section).
    # The packet appears twice: first with level-fund adjustments, then
    # level-service only. We want the first occurrence of each school.
    # Use the midpoint of the text as a safe upper bound.
    midpoint = len(text) // 2

    for slug, header_substr, display in SCHOOLS:
        idx = text.find(header_substr)
        if idx < 0 or idx > midpoint:
            # Try again -- Athletics header doesn't include "Marblehead Public Schools -"
            # in the same way; fall through gracefully.
            continue
        # Look in the section starting at this header through a generous window.
        section_end = min(idx + 60_000, midpoint)
        # Tighten to the next school header if it's closer.
        for s2, hs2, _ in SCHOOLS:
            if s2 == slug:
                continue
            other = text.find(hs2, idx + len(header_substr))
            if 0 < other < section_end:
                section_end = other
        section = text[idx:section_end]
        totals = list(total_re.finditer(section))
        if len(totals) >= 2:
            # Take the second TOTAL -- level fund (what gets appropriated).
            m = totals[1]
        elif totals:
            m = totals[0]
        else:
            continue
        fy26 = int(m.group(1).replace(",", ""))
        fy27 = int(m.group(2).replace(",", ""))
        # change_dollars from the regex may have wrong sign (parens notation);
        # always recompute from the actual dollar diff for accuracy.
        change = fy27 - fy26
        pct = (change / fy26) if fy26 else 0.0
        rows.append({
            "id": slug,
            "level": "department",
            "parent_id": "schools",
            "function": "schools",
            "department": slug,
            "description": display,
            "spend_type": None,
            "fy25_budget": None, "fy25_actual": None,
            "fy26_budget": fy26,
            "fy27_proposed": fy27,
            "change_dollars": change,
            "change_pct": pct,
            "source_ref": {"doc": "fy27_school_packet"},
        })

        # Second pass: extract Munis-level line items from this school's section.
        # Only look in the section up to (but not including) the ADDITIONS block
        # or the second TOTAL (level-fund adjusted total), so we don't double-count.
        # Find where the first TOTAL is and stop there.
        additions_idx = section.find("ADDITIONS")
        if additions_idx > 0:
            line_search_end = additions_idx
        elif totals:
            line_search_end = totals[0].start()
        else:
            line_search_end = len(section)
        line_section = section[:line_search_end]

        seen_ids: set[str] = set()
        for lm in _MUNIS_LINE_RE.finditer(line_section):
            descr = lm.group(1).strip().rstrip("*").strip()
            org = lm.group(2)
            obj = lm.group(3)
            proj = lm.group(4)
            fy26_raw = lm.group(5)
            fy27_raw = lm.group(6)
            chg_raw = lm.group(7)

            # Skip summary/subtotal lines.
            if _SKIP_DESCR_RE.match(descr):
                continue

            # Skip header rows (no numeric codes that look like Munis ORG).
            # Description should not be a column header keyword.
            if descr.lower() in {"org", "obj", "proj", "budget", "level fund",
                                  "munis", "fy26", "fy27"}:
                continue

            line_id = f"school_{slug.replace('school_', '')}_{org}_{obj}_{proj}"
            # Deduplicate within the section (some items appear across page breaks).
            if line_id in seen_ids:
                continue
            seen_ids.add(line_id)

            fy26_val = int(fy26_raw.replace(",", ""))
            fy27_val = int(fy27_raw.replace(",", ""))
            if chg_raw.startswith("(") and chg_raw.endswith(")"):
                chg_val = -int(chg_raw[1:-1].replace(",", ""))
            else:
                chg_val = int(chg_raw.replace(",", ""))
            # Recompute from actual values for accuracy.
            chg_val = fy27_val - fy26_val
            chg_pct = (chg_val / fy26_val) if fy26_val else 0.0

            rows.append({
                "id": line_id,
                "level": "line",
                "parent_id": slug,
                "function": "schools",
                "department": slug,
                "description": descr,
                "spend_type": classify_spend_type(descr),
                "fy25_budget": None,
                "fy25_actual": None,
                "fy26_budget": fy26_val,
                "fy27_proposed": fy27_val,
                "change_dollars": chg_val,
                "change_pct": chg_pct,
                "source_ref": {
                    "doc": "fy27_school_packet",
                    "munis_org": org,
                    "munis_obj": obj,
                    "munis_proj": proj,
                },
            })

    return rows


# Excel "DEPARTMENT" column → existing town-side dept slug.
# Source: data/budget_source/FY27-Proposed-Budget-vs-FY26-w-Acct-Details.xlsx,
# the Town's Munis-account-level FY26-vs-FY27 worksheet (32 distinct dept
# names). The Excel typo "COMMUNTY DEVELOPMENT" is reproduced verbatim from
# the source. Snow & Ice rolls into the Public Works ops department in our
# tree because the function-level rollup already includes Snow Removal as a
# line under "Public Works (Highway, Tree, Drains)".
_EXCEL_DEPT_TO_SLUG = {
    "MODERATOR":                    "moderator",
    "SELECT BOARD":                 "select_board",
    "FINANCE COMMITTEE":            "finance_committee",
    "RESERVE FUND":                 "reserve_fund",
    "FINANCE":                      "finance",
    "ASSESSOR":                     "assessor",
    "TOWN COUNSEL":                 "town_counsel",
    "HUMAN RESOURCES":              "human_resources",
    "PARKING CLERK":                "parking_clerk",
    "TOWN CLERK":                   "town_clerk",
    "ELECTIONS":                    "election_registration",
    "PLANNING BOARD":               "planning_board",
    "COMMUNTY DEVELOPMENT":         "community_development",
    "PUBLIC BUILDINGS":             "public_buildings",
    "POLICE":                       "police",
    "FIRE":                         "fire",
    "BUILDING INSPECTIONS":         "building_inspection",
    "SEALER OF WEIGHTS & MEASURES": "sealer_weights_measures",
    "ANIMAL INSPECTOR":             "animal_inspector",
    "PUBLIC WORKS":                 "public_works_ops",
    "SNOW & ICE":                   "public_works_ops",
    "WASTE COLLECTION":             "waste_collection",
    "CURBSIDE COLLECTION":          "curbside_collection",
    "CEMETERY":                     "cemetery",
    "HEALTH":                       "health",
    "COUNCIL ON AGING":             "council_on_aging",
    "VETERANS SERVICES":            "veterans_benefits",
    "LIBRARY":                      "library",
    "RECREATION & PARKS":           "rec_park",
    "MEMORIAL & VETERANS DAY":      "memorial_veterans_day",
    "OTHER GENERAL GOVERNMENT":     "other_general_government_dept",
    "DEBT EXCLUSIONS":              "debt_service",
}


def parse_excel_account_details(xlsx_path: Path) -> dict[str, list[dict]]:
    """Parse the FY27-Proposed-Budget-vs-FY26-w-Acct-Details.xlsx workbook
    into per-department line-item rows.

    The Excel is the most current source the Town publishes for general-fund
    spending: every Munis account code (~600 lines) with FY26 budget, FY27
    proposed, $ change, and % change. It does not contain FY25 actuals or
    enterprise funds; for those the TXT budget book remains authoritative.

    Returns a dict mapping dept slug → list of line-row dicts ready to splice
    into the main row list.
    """
    try:
        import openpyxl  # imported lazily so non-build code paths don't need it
    except ImportError as e:
        raise ImportError(
            "openpyxl is required to parse the Excel account-detail workbook. "
            "Install with: pip install openpyxl"
        ) from e

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    by_dept: dict[str, list[dict]] = {}
    seen_ids: set[str] = set()

    for r in range(2, ws.max_row + 1):
        dept_name = ws.cell(r, 1).value
        org = ws.cell(r, 2).value
        obj = ws.cell(r, 3).value
        descr = ws.cell(r, 4).value
        fy26 = ws.cell(r, 5).value
        fy27 = ws.cell(r, 6).value

        # Skip blank rows, header repeats, and section subtotals (which
        # have null department + null description but populated amounts).
        if not dept_name or dept_name == "DEPARTMENT" or not descr:
            continue

        dept_name = dept_name.strip()
        slug = _EXCEL_DEPT_TO_SLUG.get(dept_name)
        if slug is None:
            # Unknown dept; skip rather than guess. Tests will catch coverage gaps.
            continue

        descr = descr.strip()
        org = str(org).strip() if org is not None else ""
        obj = str(obj).strip() if obj is not None else ""

        fy26_v = int(fy26) if fy26 is not None else 0
        fy27_v = int(fy27) if fy27 is not None else 0
        change = fy27_v - fy26_v
        pct = (change / fy26_v) if fy26_v else 0.0

        # Use Munis ORG+OBJ for the line ID. A given dept can have the same
        # OBJ code under multiple ORG codes (e.g. salary lines for different
        # divisions), so include both.
        line_id = f"acct_{slug}_{org}_{obj}"
        if line_id in seen_ids:
            # Duplicate row -- shouldn't happen in clean source, but guard.
            continue
        seen_ids.add(line_id)

        by_dept.setdefault(slug, []).append({
            "id": line_id,
            "level": "line",
            "parent_id": slug,
            "function": None,  # filled in when spliced (parent dept knows its function)
            "department": slug,
            "description": descr,
            "spend_type": classify_spend_type(descr, obj),
            "fy25_budget": None,
            "fy25_actual": None,
            "fy26_budget": fy26_v,
            "fy27_proposed": fy27_v,
            "change_dollars": change,
            "change_pct": pct,
            "source_ref": {
                "doc": "fy27_account_details",
                "munis_org": org,
                "munis_obj": obj,
            },
        })

    return by_dept


def merge_excel_into_rows(rows: list[dict],
                          excel_lines_by_dept: dict[str, list[dict]]) -> list[dict]:
    """Replace TXT-derived line-rows for any town dept that the Excel covers,
    using the Excel's finer Munis-level lines instead. Recompute the dept
    rollups from the Excel sums for FY26/FY27 (preserving FY25 from TXT, since
    Excel doesn't include it).

    Enterprise funds (sewer, water, harbor) and Schools cost-centers are not
    covered by the Excel and pass through untouched.
    """
    covered_slugs = set(excel_lines_by_dept.keys())
    by_id = {r["id"]: r for r in rows}

    # Drop existing line rows for any dept the Excel covers.
    rows = [
        r for r in rows
        if not (r["level"] == "line" and r.get("department") in covered_slugs)
    ]

    # Splice in Excel lines and recompute dept rollups.
    for slug, lines in excel_lines_by_dept.items():
        dept = by_id.get(slug)
        if dept is None:
            # Excel mentions a dept slug not in the TXT-built rows. Skip
            # rather than fabricate a parent we don't have function info for.
            continue
        fn = dept.get("function")
        sum_fy26 = sum(l["fy26_budget"] for l in lines)
        sum_fy27 = sum(l["fy27_proposed"] for l in lines)
        for l in lines:
            l["function"] = fn  # link to function via the dept
            rows.append(l)
        # Update dept rollup from Excel sums; Excel is the more current source.
        # Preserve FY25 columns from TXT since Excel doesn't include them.
        dept["fy26_budget"] = sum_fy26
        dept["fy27_proposed"] = sum_fy27
        dept["change_dollars"] = sum_fy27 - sum_fy26
        dept["change_pct"] = (dept["change_dollars"] / sum_fy26) if sum_fy26 else 0.0

    return rows


def build_lookup(rows: list[dict]) -> dict:
    """Build a {slug: display_name} lookup helper."""
    return {r["id"]: r.get("description") for r in rows}


def build_meta(rows: list[dict]) -> dict:
    by_id = {r["id"]: r for r in rows}
    gf = by_id.get("total_general_fund", {})
    tb = by_id.get("total_budgets", {})
    from datetime import datetime, timezone
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source_doc": "FY27 Proposed Budget — No Override (function/dept totals + FY25 actuals)",
        "source_pdf_url": "https://www.marbleheadma.gov/finance-department/files/fy27-proposed-budget-no-override",
        "account_detail_doc": "FY27 Proposed Budget vs FY26 with Account Details",
        "account_detail_url": "https://marbleheadma.gov/wp-content/uploads/2026/05/FY27-Proposed-Budget-vs-FY26-w-Acct-Details.xlsx",
        "school_packet_url": "https://www.marbleheadschools.org/school-committee/files/fy27-proposed-budget-packet",
        "history_source": "DOR Schedule A function-level expenditures FY02-FY24",
        "total_general_fund": gf.get("fy27_proposed"),
        "total_with_enterprise": tb.get("fy27_proposed"),
    }


def update_general_fund_total(rows: list[dict]) -> None:
    """After Excel splice, the general-fund grand total may differ from the
    TXT-derived value because the TXT pipeline missed a few Munis-level
    lines (e.g. Transfer to Workers Comp) that the Excel includes. Recompute
    the GF total as the sum of GF function rows so the meta tile and the
    grand-total row reconcile to the per-dept tree.
    """
    by_id = {r["id"]: r for r in rows}
    gf_total = by_id.get("total_general_fund")
    if gf_total is None:
        return
    GF_FUNCTIONS = [
        "general_government", "public_safety", "public_works",
        "human_services", "schools", "culture_recreation",
        "other_general_government",
    ]
    # Sum the *function* rows (Town Meeting vote totals) + debt_service.
    # The budget book treats "TOTAL DEBT SERVICE" as a parallel subtotal
    # under Other General Government -- it is not part of the OGG function
    # vote total but does count toward TOTAL GENERAL FUND ACCOUNTS.
    sum_fy26 = sum(by_id[fn]["fy26_budget"] for fn in GF_FUNCTIONS if fn in by_id)
    sum_fy27 = sum(by_id[fn]["fy27_proposed"] for fn in GF_FUNCTIONS if fn in by_id)
    if "debt_service" in by_id:
        sum_fy26 += by_id["debt_service"]["fy26_budget"]
        sum_fy27 += by_id["debt_service"]["fy27_proposed"]
    gf_total["fy26_budget"] = sum_fy26
    gf_total["fy27_proposed"] = sum_fy27
    gf_total["change_dollars"] = sum_fy27 - sum_fy26
    gf_total["change_pct"] = (gf_total["change_dollars"] / sum_fy26) if sum_fy26 else 0.0

    # Total budgets = GF + enterprise functions.
    tb = by_id.get("total_budgets")
    if tb is None:
        return
    ENTERPRISE_FUNCTIONS = ["sewer_enterprise", "water_enterprise", "harbor_enterprise"]
    ent_fy26 = sum(by_id[fn]["fy26_budget"] for fn in ENTERPRISE_FUNCTIONS if fn in by_id)
    ent_fy27 = sum(by_id[fn]["fy27_proposed"] for fn in ENTERPRISE_FUNCTIONS if fn in by_id)
    tb["fy26_budget"] = sum_fy26 + ent_fy26
    tb["fy27_proposed"] = sum_fy27 + ent_fy27
    tb["change_dollars"] = tb["fy27_proposed"] - tb["fy26_budget"]
    tb["change_pct"] = (tb["change_dollars"] / tb["fy26_budget"]) if tb["fy26_budget"] else 0.0


if __name__ == "__main__":
    text = (DATA / "FY27_Proposed_Budget_No_Override.txt").read_text()
    rows = parse_budget_book(text)
    attach_function_history(rows)
    rows.extend(parse_school_packet())
    # The town book's "School Department" wrapper department and its single
    # "101 Schools" appropriation line are redundant with the Schools function
    # row (also $47.62M) and the school packet's per-school cost-center
    # breakdown. Drop them so the schools tree reads cleanly.
    SCHOOLS_NOISE_IDS = {"schools_dept_wrapper", "line_101"}
    rows = [r for r in rows if r["id"] not in SCHOOLS_NOISE_IDS]

    # Splice in Munis-account-level line items from the Excel workbook.
    # Excel covers all general-fund Town departments (no schools, no enterprise);
    # for each dept it covers we replace the rolled-up "Salaries / Expense"
    # lines from TXT with one line per Munis account code.
    xlsx_path = DATA / "budget_source" / "FY27-Proposed-Budget-vs-FY26-w-Acct-Details.xlsx"
    if xlsx_path.exists():
        excel_lines = parse_excel_account_details(xlsx_path)
        rows = merge_excel_into_rows(rows, excel_lines)
        update_general_fund_total(rows)
    else:
        print(f"WARN: {xlsx_path} not found; skipping account-detail splice. "
              f"Output will use TXT-only Salaries/Expense rollups.")

    out = {"meta": build_meta(rows), "rows": rows}
    (DATA / "town_budget_FY27.json").write_text(json.dumps(out, indent=2) + "\n")
    (DATA / "town_budget_FY27_lookup.json").write_text(
        json.dumps(build_lookup(rows), indent=2) + "\n"
    )
    print(f"Wrote {len(rows)} rows to town_budget_FY27.json")
    print(f"Total General Fund: ${out['meta']['total_general_fund']:,}")
    print(f"Total Budgets: ${out['meta']['total_with_enterprise']:,}")
